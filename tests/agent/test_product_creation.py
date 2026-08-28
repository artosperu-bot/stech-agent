from __future__ import annotations

from decimal import Decimal

import pytest
from openpyxl import load_workbook

from stech_agent.agent.product_creation import build_create_workbook, prepare_new_product
from stech_agent.catalog.reader import CatalogSnapshotData
from stech_agent.domain.models import ProductRecord


RAW_HEADERS = (
    "SKU", "Nombre del producto", "Descripcion", "Categoria", "Subcategoría", "Marca",
    "Precio", "Descuento", "Stock", "Es nuevo", "En oferta", "Recomendado", "Destacado",
    "Visible", "Estado", "Especificaciones principales (separadas por coma)",
    "Especificaciones técnicas (separadas por slash y dos puntos)", "Imagen", "Galería",
    "Regla de descuento", "Promociones",
)
CANONICAL_HEADERS = (
    "sku", "name", "description", "category", "subcategory", "brand", "price", "discount",
    "stock", "is_new", "on_offer", "recommended", "featured", "visible", "status", "main_specs",
    "technical_specs", "image", "gallery", "discount_rule", "promotions",
)


def snapshot() -> CatalogSnapshotData:
    return CatalogSnapshotData(
        raw_headers=RAW_HEADERS,
        canonical_headers=CANONICAL_HEADERS,
        products=(
            ProductRecord(sku="J1", name="JBL Uno", brand="JBL", category="Audio", subcategory="Parlantes Bluetooth", status="Activo", source_order=1),
            ProductRecord(sku="E1", name="Epson Uno", brand="EPSON", category="Impresión", subcategory="Impresoras", status="Activo", source_order=2),
        ),
        source_path="fixture.xlsx",
        checksum="fixture",
    )


def test_prepare_new_product_validates_and_canonicalizes_catalog_taxonomy():
    draft = prepare_new_product(snapshot(), {
        "sku": "  NEW-001 ",
        "name": "JBL Nuevo",
        "brand": "jbl",
        "category": "audio",
        "subcategory": "parlantes bluetooth",
        "price": "199.90",
        "stock": "5",
    })
    assert draft.sku == "NEW-001"
    assert draft.values["brand"] == "JBL"
    assert draft.values["category"] == "Audio"
    assert draft.values["subcategory"] == "Parlantes Bluetooth"
    assert draft.values["price"] == Decimal("199.90")
    assert draft.values["stock"] == 5
    assert draft.values["discount"] == Decimal("0")
    assert draft.values["visible"] is False
    assert draft.values["on_offer"] is False
    assert draft.values["recommended"] is False
    assert draft.values["featured"] is False
    assert draft.values["is_new"] is False


def test_prepare_new_product_rejects_existing_sku():
    with pytest.raises(ValueError, match="ya existe"):
        prepare_new_product(snapshot(), {
            "sku": "J1", "name": "Otro", "brand": "JBL", "category": "Audio",
            "subcategory": "Parlantes Bluetooth", "price": 1, "stock": 0,
        })


def test_prepare_new_product_rejects_invalid_category_subcategory_pair():
    with pytest.raises(ValueError, match="Subcategoría"):
        prepare_new_product(snapshot(), {
            "sku": "NEW-002", "name": "Error", "brand": "JBL", "category": "Audio",
            "subcategory": "Impresoras", "price": 1, "stock": 0,
        })


def test_prepare_new_product_requires_core_fields():
    with pytest.raises(ValueError, match="Faltan campos obligatorios"):
        prepare_new_product(snapshot(), {"sku": "NEW-003", "name": "Incompleto"})


def test_build_create_workbook_preserves_export_schema_and_sku_as_text(tmp_path):
    snap = snapshot()
    draft = prepare_new_product(snap, {
        "sku": "0667C999",
        "name": "Producto Nuevo",
        "description": "Descripción de prueba",
        "brand": "JBL",
        "category": "Audio",
        "subcategory": "Parlantes Bluetooth",
        "price": "10.50",
        "stock": 2,
        "main_specs": "Bluetooth, portátil, compacto",
    })
    path = tmp_path / "nuevo.xlsx"
    receipt = build_create_workbook(snap, draft, path)

    assert receipt.path == path
    assert receipt.sku == "0667C999"
    wb = load_workbook(path)
    ws = wb.active
    assert tuple(cell.value for cell in ws[1]) == RAW_HEADERS
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "0667C999"
    assert ws.cell(2, 1).number_format == "@"
    assert ws.cell(2, 2).value == "Producto Nuevo"
    assert ws.cell(2, 7).value == 10.5
    assert ws.cell(2, 9).value == 2
    assert ws.cell(2, 14).value == "No"
    wb.close()
