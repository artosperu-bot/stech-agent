from __future__ import annotations

import json

from openpyxl import load_workbook

from stech_agent.catalog.reader import CatalogSnapshotData
from stech_agent.db.connection import AgentDatabase
from stech_agent.db.migrations import migrate
from stech_agent.db.repositories import CatalogRepository
from stech_agent.domain.models import ProductRecord
from stech_agent.seo.batches import SeoBatchRepository
from stech_agent.seo.staging import export_batch_staging


def test_staging_workbook_projects_sqlite_state_one_row_per_sku(tmp_path):
    db = AgentDatabase(tmp_path / "agent.sqlite3")
    migrate(db)
    CatalogRepository(db).save_snapshot(CatalogSnapshotData(
        raw_headers=("SKU", "Nombre del producto", "Marca", "Categoria", "Subcategoría"),
        canonical_headers=("sku", "name", "brand", "category", "subcategory"),
        products=(
            ProductRecord(sku="A", name="Producto A", brand="JBL", category="Audio", subcategory="Parlantes", source_order=1),
            ProductRecord(sku="B", name="Producto B", brand="EPSON", category="Impresión", subcategory="Impresoras", source_order=2),
        ),
        source_path="fixture.xlsx",
        checksum="fixture",
    ))
    repo = SeoBatchRepository(db)
    batch_id = repo.create(session_id=None, skus=["A", "B"], scope={"all": True}, publish=False)
    items = repo.list_items(batch_id)
    with db.transaction(immediate=True) as con:
        con.execute(
            "INSERT INTO seo_research(batch_item_id,status,facts_json,sources_json,prompt_id,prompt_version,prompt_hash,provider_id) VALUES (?,?,?,?,?,?,?,?)",
            (items[0].id, "RESEARCHED", json.dumps({"modelo":"Charge"}), json.dumps(["https://jbl.com"]), "SEO_PRODUCTO_STECH_V1", "1", "abc", "edge-chatgpt"),
        )
        con.execute(
            "INSERT INTO seo_proposals(batch_item_id,current_seo_json,generated_json,proposed_patch_json,qa_status,qa_notes_json) VALUES (?,?,?,?,?,?)",
            (items[0].id, json.dumps({"seo_title":""}), json.dumps({"titulo_seo":"Nuevo"}), json.dumps({"seo_title":"Nuevo"}), "READY", json.dumps(["ok"])),
        )
        con.execute("UPDATE seo_batch_items SET state='READY' WHERE id=?", (items[0].id,))

    path = export_batch_staging(db, batch_id, tmp_path / "staging.xlsx")

    assert path.exists()
    wb = load_workbook(path, data_only=True)
    ws = wb["SEO Batch"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:6] == ["Batch ID", "SKU", "Producto", "Marca", "Categoría", "Subcategoría"]
    assert ws.max_row == 3
    rows = {ws.cell(r, 2).value: [ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, 4)}
    assert rows["A"][headers.index("Estado") ] == "READY"
    assert rows["A"][headers.index("Research Sources") ] == "https://jbl.com"
    assert rows["A"][headers.index("QA Status") ] == "READY"
    assert rows["B"][headers.index("Estado") ] == "RESEARCH_PENDING"
