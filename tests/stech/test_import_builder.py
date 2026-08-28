from openpyxl import load_workbook, Workbook
import pytest

from stech_agent.catalog.import_builder import build_import_workbook
from stech_agent.catalog.reader import read_items_export
from stech_agent.domain.models import FieldPatch


def make_source(path, rows):
    wb = Workbook(); ws = wb.active
    ws.append(["SKU", "Nombre del producto", "Marca", "Precio", "Stock", "En oferta", "Visible", "Proveedor Nuevo"])
    for r in rows: ws.append(r)
    wb.save(path); return path


def test_builder_preserves_all_source_values_except_explicit_patch(tmp_path):
    src = make_source(tmp_path / "source.xlsx", [["0667C001", "Canon", "CANON", 119.0, 3, "No", "Si", "X"]])
    snap = read_items_export(src); out = tmp_path / "import.xlsx"
    receipt = build_import_workbook(snap, {"0667C001": FieldPatch({"on_offer": True})}, out)
    assert receipt.skus == ("0667C001",)
    wb = load_workbook(out, data_only=False); ws = wb.active
    row = dict(zip([c.value for c in ws[1]], [c.value for c in ws[2]]))
    assert row["SKU"] == "0667C001" and ws["A2"].number_format == "@"
    assert row["Nombre del producto"] == "Canon" and row["Precio"] == 119.0 and row["Stock"] == 3
    assert row["En oferta"] == "Si" and row["Visible"] == "Si" and row["Proveedor Nuevo"] == "X"


def test_builder_does_not_include_unselected_sku(tmp_path):
    src = make_source(tmp_path / "source.xlsx", [["A", "A", "X", 1, 1, "No", "Si", ""], ["B", "B", "X", 2, 2, "No", "Si", ""]])
    snap = read_items_export(src); out = tmp_path / "import.xlsx"
    build_import_workbook(snap, {"B": FieldPatch({"on_offer": True})}, out)
    ws = load_workbook(out, read_only=True).active
    assert ws.max_row == 2 and ws["A2"].value == "B"


def test_builder_refuses_field_without_export_header(tmp_path):
    src = make_source(tmp_path / "source.xlsx", [["A", "A", "X", 1, 1, "No", "Si", ""]]); snap = read_items_export(src)
    with pytest.raises(ValueError, match="sin columna de exportación"): build_import_workbook(snap, {"A": FieldPatch({"description": "x"})}, tmp_path / "x.xlsx")


def test_builder_refuses_ambiguous_sku(tmp_path):
    src = make_source(tmp_path / "source.xlsx", [["A", "A", "X", 1, 1, "No", "Si", ""], ["A", "A", "X", 1, 9, "No", "Si", ""]]); snap = read_items_export(src)
    assert snap.products[0].ambiguous
    with pytest.raises(ValueError, match="ambiguo"): build_import_workbook(snap, {"A": FieldPatch({"on_offer": True})}, tmp_path / "x.xlsx")
