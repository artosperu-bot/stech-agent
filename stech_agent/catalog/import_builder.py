from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from stech_agent.catalog.reader import CatalogSnapshotData
from stech_agent.domain.fields import coerce_field
from stech_agent.domain.models import FieldPatch


@dataclass(frozen=True, slots=True)
class ImportBuildReceipt:
    path: Path
    skus: tuple[str, ...]
    fields: frozenset[str]


def _export_value(field: str, value: Any) -> Any:
    value = coerce_field(field, value)
    if isinstance(value, bool):
        return "Si" if value else "No"
    if isinstance(value, Decimal):
        return float(value)
    return value


def build_import_workbook(
    snapshot: CatalogSnapshotData,
    patches: dict[str, FieldPatch],
    output_path: str | Path,
) -> ImportBuildReceipt:
    if not patches:
        raise ValueError("No hay cambios para importar")
    header_for_field = {
        canonical: raw
        for raw, canonical in zip(snapshot.raw_headers, snapshot.canonical_headers)
        if not canonical.startswith("extra:")
    }
    products = {p.sku: p for p in snapshot.products}
    requested_fields: set[str] = set()
    for sku, patch in patches.items():
        if sku not in products:
            raise ValueError(f"SKU no existe en snapshot: {sku}")
        product = products[sku]
        if product.ambiguous:
            raise ValueError(f"SKU ambiguo {sku}; conflictos: {', '.join(sorted(product.conflict_fields))}")
        for field in patch.fields:
            if field == "sku":
                raise ValueError("El SKU no puede modificarse")
            if field not in header_for_field:
                raise ValueError(f"Campo {field!r} sin columna de exportación conocida")
            requested_fields.add(field)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(list(snapshot.raw_headers))
    for sku, patch in patches.items():
        product = products[sku]
        row = [product.source.get(header) for header in snapshot.raw_headers]
        for field, value in patch.values.items():
            raw_header = header_for_field[field]
            col_idx = snapshot.raw_headers.index(raw_header)
            row[col_idx] = _export_value(field, value)
        ws.append(row)
        sku_col = snapshot.canonical_headers.index("sku") + 1
        cell = ws.cell(ws.max_row, sku_col)
        cell.value = str(sku)
        cell.number_format = "@"
    wb.save(output_path)
    return ImportBuildReceipt(path=output_path, skus=tuple(patches.keys()), fields=frozenset(requested_fields))
