from __future__ import annotations

from dataclasses import dataclass, replace
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from stech_agent.domain.fields import coerce_field, resolve_header
from stech_agent.domain.models import ProductRecord


@dataclass(frozen=True, slots=True)
class CatalogSnapshotData:
    raw_headers: tuple[str, ...]
    canonical_headers: tuple[str, ...]
    products: tuple[ProductRecord, ...]
    source_path: str
    checksum: str

    @property
    def schema_headers(self) -> tuple[str, ...]:
        return self.canonical_headers


def _checksum(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_items_export(path: str | Path) -> CatalogSnapshotData:
    source_path = Path(path)
    wb = load_workbook(source_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        raw_header_values = next(rows)
    except StopIteration as exc:
        raise ValueError("Excel vacío") from exc

    raw_headers = tuple("" if h is None else str(h).strip() for h in raw_header_values)
    canonical_headers = tuple(resolve_header(h) for h in raw_headers)
    if "sku" not in canonical_headers:
        raise ValueError("No se encontró la columna SKU")

    products: list[ProductRecord] = []
    seen: dict[str, str] = {}
    index_by_sku: dict[str, int] = {}

    for source_order, row in enumerate(rows, start=1):
        source: dict[str, Any] = {header: value for header, value in zip(raw_headers, row)}
        canonical: dict[str, Any] = {}
        extra: dict[str, Any] = {}
        for key, value in zip(canonical_headers, row):
            coerced = coerce_field(key, value)
            if key.startswith("extra:"):
                extra[key] = coerced
            else:
                canonical[key] = coerced

        sku = str(canonical.get("sku") or "").strip()
        if not sku:
            continue
        name = str(canonical.get("name") or "").strip()
        if sku in seen and seen[sku] != name:
            raise ValueError(f"SKU duplicado con nombre conflictivo: {sku}")
        if sku in seen:
            idx = index_by_sku[sku]
            current = products[idx]
            differing: set[str] = set(current.conflict_fields)
            for key, value in canonical.items():
                if key in {"sku", "name"} or key.startswith("extra:"):
                    continue
                if hasattr(current, key) and getattr(current, key) != value:
                    differing.add(key)
            for key, value in extra.items():
                if current.extra.get(key) != value:
                    differing.add(key)
            products[idx] = replace(
                current,
                duplicate_sources=current.duplicate_sources + (source,),
                conflict_fields=frozenset(differing),
            )
            continue

        seen[sku] = name
        index_by_sku[sku] = len(products)
        kwargs = {k: v for k, v in canonical.items() if k in ProductRecord.__dataclass_fields__}
        products.append(ProductRecord(**kwargs, source=source, extra=extra, source_order=source_order))

    wb.close()
    return CatalogSnapshotData(
        raw_headers=raw_headers,
        canonical_headers=canonical_headers,
        products=tuple(products),
        source_path=str(source_path),
        checksum=_checksum(source_path),
    )
